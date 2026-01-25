"""
Generador de Reportes Mensuales en Excel
Formato exacto como CONRAD_CENTRAL_ENERO_2026
"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from typing import List, Dict
import calendar

class ReporteConradExcel:
    """
    Genera reportes mensuales con formato profesional
    Una hoja por día del mes
    """
    
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Remover hoja default
        
        # Colores corporativos
        self.COLOR_HEADER = "FFA4C2F4"  # Azul claro headers
        self.COLOR_TITULO = "FFD9D9D9"  # Gris claro título
        self.COLOR_TOTALES = "FFFDE9D9"  # Naranja claro totales
        
        # Fuentes
        self.FONT_HEADER = Font(name='Calibri', size=11, bold=True)
        self.FONT_TITULO = Font(name='Calibri', size=12, bold=True)
        self.FONT_DATOS = Font(name='Arial', size=10)
        self.FONT_TOTALES = Font(name='Calibri', size=11, bold=True)
    
    def crear_reporte_mensual(self, mes: int, anio: int, consultas_por_dia: Dict[int, List[Dict]]):
        """
        Crea un reporte mensual completo
        
        Args:
            mes: Número del mes (1-12)
            anio: Año (ej: 2026)
            consultas_por_dia: {dia: [consultas]}
        """
        # Obtener días del mes
        dias_mes = calendar.monthrange(anio, mes)[1]
        
        # Crear una hoja por cada día
        for dia in range(1, dias_mes + 1):
            fecha = datetime(anio, mes, dia)
            fecha_str = fecha.strftime("%d%m%y")  # Formato: 010126
            
            # Crear hoja
            ws = self.wb.create_sheet(title=fecha_str)
            
            # Consultas del día (si existen)
            consultas = consultas_por_dia.get(dia, [])
            
            # Generar contenido de la hoja
            self._crear_hoja_diaria(ws, fecha, consultas)
        
        return self.wb
    
    def _crear_hoja_diaria(self, ws, fecha: datetime, consultas: List[Dict]):
        """Genera el contenido de una hoja diaria"""
        
        # ===== CONFIGURAR ANCHOS DE COLUMNA =====
        anchos = {
            'A': 4.75,
            'B': 24.0,   # NOMBRE
            'C': 6.75,   # EDAD
            'D': 13.0,   # NO. FACTURA
            'E': 25.38,  # ESTUDIO
            'F': 13.0,   # MEDICO
            'G': 13.0,   # PRECIO SOCIAL
            'H': 13.0,   # RX
            'I': 13.0,   # USG
            'J': 8.63,   # TAC
            'K': 13.0,   # EKG/PAP/LABS
            'L': 13.0,   # ESTADO CUENTA
            'M': 5.75,   # TIPO
            'N': 13.0
        }
        
        for col, ancho in anchos.items():
            ws.column_dimensions[col].width = ancho
        
        # ===== FILA 1: TÍTULO CON FECHA =====
        fecha_formateada = fecha.strftime("%d/%m/%Y")
        ws['B1'] = f"FECHA: {fecha_formateada}"
        ws['D1'] = "CONRAD CENTRAL"
        
        # Estilo título
        for cell in [ws['B1'], ws['D1']]:
            cell.font = self.FONT_TITULO
            cell.fill = PatternFill(start_color=self.COLOR_TITULO, 
                                   end_color=self.COLOR_TITULO, 
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Merge cells del título
        ws.merge_cells('B1:C1')
        ws.merge_cells('D1:G1')
        
        # ===== FILA 2: HEADERS =====
        headers = [
            ('B2', 'NOMBRE DEL PACIENTE'),
            ('C2', 'EDAD'),
            ('D2', 'NO. FACTURA'),
            ('E2', 'ESTUDIO'),
            ('F2', 'MEDICO REFERENTE'),
            ('G2', 'PRECIO SOCIAL'),
            ('H2', 'RX'),
            ('I2', 'USG'),
            ('J2', 'TAC'),
            ('K2', 'EKG/PAP/LABS'),
            ('L2', 'ESTADO DE CUENTA'),
            ('M2', 'TIPO')
        ]
        
        for cell_ref, texto in headers:
            cell = ws[cell_ref]
            cell.value = texto
            cell.font = self.FONT_HEADER
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, 
                                   end_color=self.COLOR_HEADER, 
                                   fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        ws.row_dimensions[2].height = 30  # Altura del header
        
        # ===== FILAS 3+: DATOS DE PACIENTES =====
        fila_actual = 3
        
        for idx, consulta in enumerate(consultas, 1):
            # Número correlativo
            ws[f'A{fila_actual}'] = idx
            ws[f'A{fila_actual}'].font = self.FONT_DATOS
            ws[f'A{fila_actual}'].alignment = Alignment(horizontal='center')
            
            # Nombre paciente
            ws[f'B{fila_actual}'] = consulta.get('nombre_paciente', '').upper()
            ws[f'B{fila_actual}'].font = self.FONT_DATOS
            
            # Edad
            ws[f'C{fila_actual}'] = consulta.get('edad', '')
            ws[f'C{fila_actual}'].font = self.FONT_DATOS
            ws[f'C{fila_actual}'].alignment = Alignment(horizontal='center')
            
            # No. Factura
            ws[f'D{fila_actual}'] = consulta.get('numero_factura', '')
            ws[f'D{fila_actual}'].font = self.FONT_DATOS
            ws[f'D{fila_actual}'].alignment = Alignment(horizontal='center')
            
            # Estudio
            ws[f'E{fila_actual}'] = consulta.get('estudio', '').upper()
            ws[f'E{fila_actual}'].font = self.FONT_DATOS
            
            # Médico referente
            ws[f'F{fila_actual}'] = consulta.get('medico', 'TRATANTE').upper()
            ws[f'F{fila_actual}'].font = self.FONT_DATOS
            
            # Precio Social
            if consulta.get('tipo_cobro') == 'social':
                ws[f'G{fila_actual}'] = consulta.get('precio', 0)
                ws[f'G{fila_actual}'].number_format = '#,##0.00'
            
            # Distribución por categoría
            categoria = consulta.get('categoria', 'RX')
            precio = consulta.get('precio', 0)
            
            if categoria == 'RX':
                ws[f'H{fila_actual}'] = precio
                ws[f'H{fila_actual}'].number_format = '#,##0.00'
            elif categoria == 'USG':
                ws[f'I{fila_actual}'] = precio
                ws[f'I{fila_actual}'].number_format = '#,##0.00'
            elif categoria == 'TAC':
                ws[f'J{fila_actual}'] = precio
                ws[f'J{fila_actual}'].number_format = '#,##0.00'
            else:  # EKG/PAP/LABS
                ws[f'K{fila_actual}'] = precio
                ws[f'K{fila_actual}'].number_format = '#,##0.00'
            
            # Estado de cuenta
            if consulta.get('forma_pago') == 'estado_cuenta':
                ws[f'L{fila_actual}'] = precio
                ws[f'L{fila_actual}'].number_format = '#,##0.00'
            
            # Tipo (P = Particular)
            ws[f'M{fila_actual}'] = 'P'
            ws[f'M{fila_actual}'].font = self.FONT_DATOS
            ws[f'M{fila_actual}'].alignment = Alignment(horizontal='center')
            
            fila_actual += 1
        
        # Si no hay consultas, dejar al menos 3 filas vacías
        if len(consultas) == 0:
            fila_actual = 6
        
        # ===== SECCIÓN DE TOTALES =====
        fila_totales = max(fila_actual + 2, 8)
        ultima_fila_datos = fila_actual - 1
        
        # Labels y fórmulas de totales
        totales = [
            ('EFECTIVO', f'=SUM(G{fila_totales+4}-G{fila_totales+3}-G{fila_totales+2}-G{fila_totales+1})'),
            ('DEPOSITADO', 0),
            ('TARJETA', 0),
            ('ESTADO DE CUENTA', f'=SUM(L3:L{ultima_fila_datos})'),
            ('TOTAL GENERADO', f'=SUM(H3:K{ultima_fila_datos})')
        ]
        
        for idx, (label, valor) in enumerate(totales):
            fila = fila_totales + idx
            
            # Label
            ws[f'F{fila}'] = label
            ws[f'F{fila}'].font = self.FONT_TOTALES
            ws[f'F{fila}'].fill = PatternFill(start_color=self.COLOR_TOTALES, 
                                              end_color=self.COLOR_TOTALES, 
                                              fill_type="solid")
            ws[f'F{fila}'].alignment = Alignment(horizontal='right', vertical='center')
            
            # Valor
            ws[f'G{fila}'] = valor
            ws[f'G{fila}'].font = self.FONT_TOTALES
            ws[f'G{fila}'].fill = PatternFill(start_color=self.COLOR_TOTALES, 
                                              end_color=self.COLOR_TOTALES, 
                                              fill_type="solid")
            ws[f'G{fila}'].number_format = '#,##0.00'
            ws[f'G{fila}'].alignment = Alignment(horizontal='right', vertical='center')
            
            # Bordes
            for cell in [ws[f'F{fila}'], ws[f'G{fila}']]:
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
    
    def guardar(self, nombre_archivo: str):
        """Guarda el workbook"""
        self.wb.save(nombre_archivo)
        return nombre_archivo


# ===== FUNCIÓN HELPER PARA USAR EN EL SISTEMA =====

def generar_reporte_mensual_conrad(mes: int, anio: int, consultas_por_dia: Dict[int, List[Dict]], 
                                   nombre_archivo: str = None) -> str:
    """
    Genera un reporte mensual de CONRAD en formato Excel
    
    Args:
        mes: 1-12
        anio: Año completo (ej: 2026)
        consultas_por_dia: {
            1: [
                {
                    'nombre_paciente': 'JUAN PEREZ',
                    'edad': 25,
                    'numero_factura': '',
                    'estudio': 'RX TORAX',
                    'medico': 'DR. LOPEZ',
                    'tipo_cobro': 'normal',  # 'normal', 'social', 'especial'
                    'forma_pago': 'efectivo',  # 'efectivo', 'tarjeta', 'estado_cuenta'
                    'categoria': 'RX',  # 'RX', 'USG', 'TAC', 'LAB'
                    'precio': 150.00
                },
                ...
            ],
            2: [...],
            ...
        }
        nombre_archivo: Ruta donde guardar (si None, se genera automáticamente)
    
    Returns:
        Ruta del archivo generado
    """
    if nombre_archivo is None:
        nombre_mes = calendar.month_name[mes].upper()
        nombre_archivo = f"CONRAD_CENTRAL_{nombre_mes}_{anio}.xlsx"
    
    generador = ReporteConradExcel()
    generador.crear_reporte_mensual(mes, anio, consultas_por_dia)
    return generador.guardar(nombre_archivo)
